import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

# Weight function
def get_weight(cat, sub):
    if pd.isna(cat) or pd.isna(sub):
        return 0
    cat = str(cat).strip()
    sub = str(sub).strip()
    if 'small diaries' in cat.lower():
        return 300
    elif 'big diaries' in cat.lower():
        return 600
    elif 'table top calendar' in cat.lower():
        return 300
    else:
        return 0

st.title("TTD Excel Processor")

uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    file_bytes = uploaded_file.getvalue()
    
    try:
        # Use openpyxl to read structure accurately
        wb_buffer = io.BytesIO(file_bytes)
        wb = load_workbook(wb_buffer, data_only=True)
        ws = wb['Sheet5']  # Use the correct sheet name
        max_col = ws.max_column

        # Read first 3 rows accurately
        row1 = [cell.value for cell in ws[1]]
        row2 = [cell.value for cell in ws[2]]
        row3 = [cell.value for cell in ws[3]]

        # Pad to max_col if shorter
        for r in [row1, row2, row3]:
            while len(r) < max_col:
                r.append(None)

        # Read data with exact columns
        data_buffer = io.BytesIO(file_bytes)
        data_df = pd.read_excel(data_buffer, sheet_name='Sheet5', header=None, skiprows=3, usecols=range(max_col))
        data_df.columns = row1

        # Compute physical weights
        data_df['PHYSICAL_WEIGHT_GRAMS'] = data_df['Quantity'] * data_df.apply(lambda row: get_weight(row['Category'], row['Sub Category']), axis=1)

        # Prepare clean headers for matching
        clean_source = {str(h).strip().lower(): h for h in row1 if h is not None}

        # Output DF
        output_df = pd.DataFrame(index=data_df.index)

        for i, desired_header in enumerate(row3):
            if desired_header is None:
                continue

            # Special cases
            desired_clean = str(desired_header).strip().lower()
            if 'physical weight' in desired_clean:
                output_df[desired_header] = data_df['PHYSICAL_WEIGHT_GRAMS']
                continue
            if 'sender add line 1' in desired_clean:
                output_df[desired_header] = 'SALES WING OF PUBLICATIONS'
                continue
            if 'sender add line 2' in desired_clean:
                output_df[desired_header] = 'TTD PRESS COMPOUND'
                continue
            if 'sender add line 3' in desired_clean:
                output_df[desired_header] = 'Tirupati - 517507'
                continue

            # Mapping
            mapped_value = row2[i]
            mapped_str = str(mapped_value).strip().lower() if mapped_value is not None else ''
            
            if mapped_str in clean_source:
                orig_col = clean_source[mapped_str]
                output_df[desired_header] = data_df[orig_col]
            else:
                # Constant - repeat the original value
                const_val = mapped_value
                output_df[desired_header] = const_val

        # Write output
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            # Headers
            pd.DataFrame([row3]).to_excel(writer, 'Sheet1', index=False, header=False, startrow=0)
            # Data
            output_df.to_excel(writer, 'Sheet1', index=False, header=False, startrow=1)

        output_buffer.seek(0)

        st.download_button(
            label="Download Processed Excel",
            data=output_buffer,
            file_name="ttd_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("File processed!")

    except Exception as e:
        st.error(f"Error: {e}")
        import traceback
        st.code(traceback.format_exc())